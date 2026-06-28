import pandas as pd
from openpyxl.styles import Font, PatternFill

rows = []

def add(id_, text, topic, type_, notes=""):
    rows.append({"id": id_, "text": text, "topic": topic, "type": type_, "notes": notes})

# ---------------------------------------------------------------- 8 disjoint clusters, 6 docs each
clusters = {
    "cooking": [
        "Sauteing vegetables over high heat caramelizes their natural sugars quickly.",
        "A classic French baguette relies on long fermentation for its open crumb.",
        "Sushi rice is seasoned with a mix of vinegar, sugar, and salt.",
        "Braising tough cuts of meat slowly breaks down connective tissue into gelatin.",
        "Whisking egg whites to stiff peaks requires a completely grease-free bowl.",
        "Fermenting cabbage with salt produces the tangy crunch of sauerkraut.",
    ],
    "sports": [
        "A marathon runner's cadence often settles around 170-180 steps per minute.",
        "Basketball free throws are taken from a line 15 feet from the backboard.",
        "Swimmers reduce drag by streamlining their bodies off every wall turn.",
        "A soccer match consists of two 45-minute halves with added stoppage time.",
        "Tennis players use topspin to keep fast shots dipping inside the baseline.",
        "Olympic weightlifting includes the snatch and the clean and jerk.",
    ],
    "finance": [
        "Compound interest grows a balance faster the more frequently it's applied.",
        "A bond's yield moves inversely to its market price.",
        "Diversifying a portfolio across asset classes reduces unsystematic risk.",
        "An inverted yield curve has historically preceded several recessions.",
        "Dollar-cost averaging spreads purchases over time to smooth out volatility.",
        "A company's P/E ratio compares its share price to its earnings per share.",
    ],
    "travel": [
        "Shoulder season often offers lower airfare and thinner crowds at major sites.",
        "A valid passport with six months' remaining validity is required by many countries.",
        "Overnight trains in Europe can save a hotel night and travel time at once.",
        "Packing a universal power adapter avoids scrambling at the airport.",
        "Travel insurance with medical evacuation coverage matters most for remote trekking.",
        "Layovers under 90 minutes carry real risk of a missed connection.",
    ],
    "music": [
        "A major scale follows the pattern whole-whole-half-whole-whole-whole-half.",
        "Jazz improvisation often leans on the blues scale over a swung rhythm.",
        "A string quartet balances two violins, a viola, and a cello.",
        "Syncopation places accents on the off-beats rather than the downbeat.",
        "Vinyl mastering requires different EQ choices than a digital streaming master.",
        "A capo lets a guitarist transpose chord shapes without relearning fingerings.",
    ],
    "history": [
        "The printing press accelerated the spread of literacy across 15th-century Europe.",
        "The Silk Road connected trade between East Asia and the Mediterranean for centuries.",
        "The fall of the Berlin Wall in 1989 symbolized the end of the Cold War divide.",
        "Ancient Rome's aqueducts moved water across long distances using gravity alone.",
        "The Industrial Revolution shifted economies from agrarian to factory-based production.",
        "The Treaty of Westphalia in 1648 helped establish the modern concept of state sovereignty.",
    ],
    "medicine": [
        "Antibiotics target bacterial infections but are ineffective against viruses.",
        "Blood pressure readings combine systolic and diastolic measurements.",
        "Vaccines train the immune system to recognize a pathogen without causing illness.",
        "Chronic inflammation is linked to a range of long-term health conditions.",
        "Insulin resistance is a hallmark feature of type 2 diabetes.",
        "Physical therapy after surgery helps restore range of motion gradually.",
    ],
    "programming": [
        "A binary search halves the search space on every comparison.",
        "Garbage collection automatically reclaims memory no longer referenced.",
        "Recursive functions need a clear base case to avoid infinite calls.",
        "Version control lets multiple developers merge changes without overwriting work.",
        "A hash map offers near-constant time lookups for most workloads.",
        "Unit tests catch regressions before they reach production code.",
    ],
}

prefix = {"cooking": "coo", "sports": "spo", "finance": "fin", "travel": "tra",
          "music": "mus", "history": "his", "medicine": "med", "programming": "pro"}

for topic, texts in clusters.items():
    for i, text in enumerate(texts, 1):
        add(f"{prefix[topic]}_{i:02d}", text, topic, "core")

# ---------------------------------------------------------------- near-duplicates (tight pairs within a cluster)
add("dup_01", "Cooking vegetables quickly on high heat browns the natural sugars inside them.",
    "cooking", "near_duplicate", "paraphrase of coo_01 - expect near-identical embedding/highest similarity")
add("dup_02", "The more often interest compounds, the faster a balance increases over time.",
    "finance", "near_duplicate", "paraphrase of fin_01 - expect near-identical embedding/highest similarity")
add("dup_03", "Binary search cuts the remaining search space in half with each comparison.",
    "programming", "near_duplicate", "paraphrase of pro_01 - expect near-identical embedding/highest similarity")

# ---------------------------------------------------------------- bridge docs (sit between two clusters)
add("brg_01", "Sports medicine physicians manage athletes' injury recovery and conditioning programs.",
    "sports+medicine", "bridge", "expect to sit visually between the sports and medicine clusters")
add("brg_02", "Walking the old Silk Road trade routes today has become a niche heritage travel itinerary.",
    "travel+history", "bridge", "expect to sit visually between the travel and history clusters")

# ---------------------------------------------------------------- outliers (disjoint from everything)
add("out_01", "Quantum entanglement allows particle states to remain correlated across vast distances.",
    "physics", "outlier", "should NOT cluster tightly with any of the 8 topics")
add("out_02", "Volcanic ash plumes can disrupt commercial flight paths for hundreds of miles.",
    "geology", "outlier", "should NOT cluster tightly with any of the 8 topics")

df = pd.DataFrame(rows)
df.to_excel("/home/claude/chroma_practice_lab/dataset/practice_dataset.xlsx", index=False, sheet_name="documents")

# Light formatting pass
from openpyxl import load_workbook
wb = load_workbook("/home/claude/chroma_practice_lab/dataset/practice_dataset.xlsx")
ws = wb["documents"]
header_font = Font(bold=True, name="Arial", color="FFFFFF")
header_fill = PatternFill("solid", start_color="404040")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
widths = {"A": 12, "B": 80, "C": 18, "D": 16, "E": 60}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
wb.save("/home/claude/chroma_practice_lab/dataset/practice_dataset.xlsx")

print(f"Rows written: {len(df)}")
print(df["type"].value_counts())
print(df["topic"].value_counts())
